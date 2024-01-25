import openpyxl


inv_file = openpyxl.load_workbook("./automationProject/files/inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
products_under_10_inv = {}
total_inv_value_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
        
        product_id = product_list.cell(product_row, 1).value
        inventory = product_list.cell(product_row, 2).value
        price = product_list.cell(product_row, 3).value
        supplier_name = product_list.cell(product_row, 4).value
        inventory_price = product_list.cell(product_row, 5)

        if supplier_name in products_per_supplier:
            products_per_supplier[supplier_name] += 1
            total_inv_value_per_supplier[supplier_name] += inventory * price
        else:
            print(f"Adding new supplier: {supplier_name}")
            products_per_supplier[supplier_name] = 1
            total_inv_value_per_supplier[supplier_name] = inventory * price

        if inventory < 10:
            products_under_10_inv[product_id] = inventory

        # Ex. 04)
        inventory_price.value = inventory * price
        inv_file.save("./automationProject/files/inventory_update.xlsx")             
            
print(products_per_supplier)
print(products_under_10_inv)
print(total_inv_value_per_supplier)